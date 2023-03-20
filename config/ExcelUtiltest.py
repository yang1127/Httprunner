# 封装excel的方法
'''
import openpyxl
from openpyxl import load_workbook


class ParseExcel:

    def __init__(self, excelPath, sheetName):
        # 将要读取excel加载到内存
        self.wb = load_workbook(excelPath)

        # 通过工作表名获取一个工作表对象
        self.sheet = self.wb[sheetName]

        # 获取工作表中存在数据的区域最大行号
        self.maxRowNum = self.sheet.max_row

    # 读取excel数据
    def getDatasFromSheet(self):
        # 存放从表中取出的数据
        Data_List = []
        for line in list(self.sheet.rows)[1:]:
            # 遍历工作表中数据区域每一行
            tmplt = []
            tmplt.append(line[1].value)  # tmplt有多少列，写多少
            # tmplt.append(line[2].value)
            Data_List.append(tmplt)
            # print(Data_List)
        return Data_List  # [[],[]]

    # 写入excel数据
    def op_toExcel(data, fileName):
        # 打开指定excel文件，openpyxl库储存数据到excel
        workbook = openpyxl.load_workbook(fileName)

        # 选定一个worksheet
        worksheet = workbook['Sheet1']

        # 添加表头 - append每次都会生成表头
        # worksheet.append(['id', 'city'])

        for i in range(len(data[0])):
            d = data[i]['id'], data[i]['city']
            # 每次写入一行
            worksheet.append(d)
        workbook.save(fileName)

# 数据用例 - 写入
testdata = [
    {'id': 1, 'city': '西安'},
    {'id': 2, 'city': '临潼'}
]

fileName = '/Users/yangzhiqi/HPtest/testdata/exceldata.xlsx'
ParseExcel.op_toExcel(testdata, fileName)

# 执行打印读取的excel内容
if __name__ == '__main__':
    excelPath = '/Users/yangzhiqi/HPtest/testdata/exceldata.xlsx'
    sheetName = u'Sheet1'
    pe = ParseExcel(excelPath, sheetName)
    for i in pe.getDatasFromSheet():
        print(i[0])
'''